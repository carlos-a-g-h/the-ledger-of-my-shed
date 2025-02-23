#!/usr/bin/python3.9

from asyncio import to_thread as async_run_block
from datetime import (datetime,timedelta)
from pathlib import Path
from secrets import token_hex
from typing import Optional,Mapping

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from symbols_Any import (
	_LANG_EN,_LANG_ES,
	_KEY_DATE,_KEY_COMMENT,
	_KEY_TAG,
	_DIR_TEMP,
)

from symbols_assets import (
	_KEY_ASSET,_KEY_NAME,_KEY_VALUE,_KEY_HISTORY,
	# _KEY_RECORD_UID,
	_KEY_RECORD_MOD
)

from dbi_assets import dbi_assets_AssetQuery
from internals import (
	util_valid_int,
	util_valid_str,
	util_valid_date,
	util_dt_to_str,
	util_date_get_day,
)

_KEY_FTREND="follow-trend"
_KEY_ATYPE="atype"

_ExExErr="E.E. Error"
_ExExWarn="E.E. Warning"


_TL_DATE={
	_LANG_EN:"Date",
	_LANG_ES:"Fecha"
}

_TL_TAG={
	_LANG_EN:"Tag",
	_LANG_ES:"Etiqueta"
}

_TL_IGNORED={
	_LANG_EN:"IGNORED",
	_LANG_ES:"IGNORADO"
}

_ERR_TIMEFRAME={
	_LANG_EN:"Outside of the requested time frame",
	_LANG_ES:"Fuera del marco de tiempo solicitado"
}

_ERR_TREND={
	_LANG_EN:"Does not follow the requested trend",
	_LANG_ES:"No sigue la tendencia a analizar"
}

_TL_TREND={
	_LANG_EN:"Trend to analyze",
	_LANG_ES:"Tendencia a analizar"
}

_TL_TREND_U={
	_LANG_EN:"Uphill",
	_LANG_ES:"Alcista"
}

_TL_TREND_D={
	_LANG_EN:"Downhill",
	_LANG_ES:"Bajista"
}

# _TL_TF_SPEC={
# 	_LANG_EN:"Within the requested time frame",
# 	_LANG_ES:"Dentro del rango de tiempo especificado"
# }

# _TL_TF_START={
# 	_LANG_EN:"Time frame start",
# 	_LANG_ES:"Inicio del rango de tiempo"
# }

# _TL_TF_END={
# 	_LANG_EN:"Time frame end",
# 	_LANG_ES:"Fin del rango de tiempo"
# }


def util_filter_by_trend(mod:int,trend:int=0)->bool:

	if mod==0:
		return True

	if trend>0:
		return (mod>0)

	if trend<0:
		return (mod<0)

	return True

def util_get_initial_supply(
		history:Mapping,
		date_min:Optional[datetime]=None,
		the_trend:int=0,
	)->int:

	# NOTE: The initial supply is the total supply BEFORE the given index, so at index zero, the supply is zero

	supply=0

	for record_uid in history:

		record_date=util_valid_date(
			history[record_uid].get(_KEY_DATE),
			date_min=date_min
		)
		if record_date is not None:
			# skip, because it reaches or is beyond the min date
			continue

		record_mod=util_valid_int(
			history[record_uid].get(_KEY_RECORD_MOD),
		)
		if record_mod is None:
			continue

		supply=supply+record_mod

	return supply

def util_get_current_supply(
		history:Mapping,

		date_min:Optional[datetime]=None,
		date_max:Optional[datetime]=None,

		the_trend:int=0,

	)->int:

	supply=0

	for record_uid in history:

		record_date=util_valid_date(
			history[record_uid].get(_KEY_DATE),
			date_min=date_min,
			date_max=date_max,
		)
		if record_date is None:
			continue

		record_mod=util_valid_int(
			history[record_uid].get(_KEY_RECORD_MOD),
		)
		if record_mod is None:
			continue

		if not util_filter_by_trend(
			record_mod,
			the_trend
		):
			continue

		supply=supply+record_mod

	return supply

def conversion_process(

		lang:str,
		path_base:Path,

		list_of_assets:list,
		atype:int=0,
		inc_history:bool=False,

		date:Optional[datetime]=None,
		date_min_og:Optional[datetime]=None,
		date_max_og:Optional[datetime]=None,

		# Ignores all points in history that go against the trend in the intended analysis
		follow_trend:bool=False,

	)->Optional[Path]:

	path_output=path_base.joinpath(
		_DIR_TEMP
	).joinpath(
		f"assets_x{token_hex(8)}.xlsx"
	)
	path_output.parent.mkdir(
		parents=True,
		exist_ok=True
	)

	date_min=date_min_og
	date_max=date_max_og
	if isinstance(date,datetime):
		date_min=util_date_get_day(date)
		date_max=date_min+timedelta(days=1)

	date_min_str:Optional[str]=util_dt_to_str(date_min)
	date_max_str:Optional[str]=util_dt_to_str(date_max)

	has_tc=(
		# isinstance(date,datetime) or
		isinstance(date_min,datetime) or
		isinstance(date_max,datetime)
	)

	has_min=isinstance(date_min,datetime)
	has_max=isinstance(date_max,datetime)

	# has_timeframe=(has_min or has_max)

	the_trend=0
	if follow_trend and (not atype==0):
		the_trend=atype

	print("Is there a time frame?",has_tc)
	print("\tdate_min",date_min)
	print("\tdate_max",date_max)

	wb:Workbook=Workbook()
	ws:Worksheet=wb.active
	ws.title="SHLED_ASSETS"

	col_headers=[
		# Column A
		{
			_LANG_EN:"Asset ID",
			_LANG_ES:"ID del activo"
		}[lang],

		# Column B
		{
			_LANG_EN:"Name",
			_LANG_ES:"Nombre"
		}[lang],

		# Column C
		{
			_LANG_EN:"Tag",
			_LANG_ES:"Etiqueta"
		}[lang],

		# Column D
		{
			_LANG_EN:"Value",
			_LANG_ES:"Valor"
		}[lang],

		# Column E
		{
			_LANG_EN:"Initial Supply",
			_LANG_ES:"Cant. inicial"
		}[lang],

		# Column F
		{
			_LANG_EN:"Supply",
			_LANG_ES:"Cant. actual"
		}[lang],

	]

	# Column G (Optional)

	if not atype==0:

		tl={
			_LANG_EN:"Performance",
			_LANG_ES:"Desempeño"
		}[lang]

		col_headers.append(tl)

	# Appending the first row of columns
	ws.append(col_headers)
	row=1

	# history (if needed) starts AFTER the last column
	col_h_start=len(col_headers)+1

	# Each row is an asset
	for asset in list_of_assets:

		row=row+1

		asset_name=asset.get(_KEY_NAME)
		asset_id=asset.get(_KEY_ASSET)
		asset_tag=asset.get(_KEY_TAG)
		asset_value=asset.get(_KEY_VALUE)

		print("\n->",asset_id,asset_name)

		# THE FIRST 4 COLUMNS

		ws[f"A{row}"]=asset_id
		ws[f"B{row}"]=asset_name
		ws[f"C{row}"]=asset_tag
		ws[f"D{row}"]=asset_value

		if not isinstance(
			asset.get(_KEY_HISTORY),
			Mapping
		):
			print(
				f"The asset {asset_id} has no history"
			)

			ws[f"E{row}"]=0

			if not atype==0:
				ws[f"F{row}"]=0

			continue

		history_size=len(asset[_KEY_HISTORY])
		if history_size==0:
			print(
				_ExExWarn,
				f"{asset_id} has history of lengh zero"
			)

		col_pos=4
		col_pos_ok=""

		# NEXT COLUMN (E) - THE INITIAL SUPPLY

		col_pos=col_pos+1
		# col_pos_ok=util_excel_dectocol(col_pos)
		col_pos_ok=get_column_letter(col_pos)
		# sheet_cell=f"{util_excel_dectocol(col_pos)}{row}"
		sheet_cell=f"{get_column_letter(col_pos)}{row}"

		sheet_cell_isup=sheet_cell

		supply_init=util_get_initial_supply(
			asset[_KEY_HISTORY],
			date_min,
			the_trend=the_trend
		)

		ws[sheet_cell]=supply_init

		# NEXT COLUMN (F) - THE SUPPLY

		col_pos=col_pos+1
		col_pos_ok=get_column_letter(col_pos)
		sheet_cell=f"{get_column_letter(col_pos)}{row}"
		sheet_cell_csup=sheet_cell

		if inc_history:

			if history_size==0:
				ws[sheet_cell]=0

			if history_size>0:
				tl=f"={sheet_cell_isup}"

				if history_size>1:
					tl=(
						f"{tl} + SUM("
							f"{get_column_letter(col_h_start)}{row}:{get_column_letter(col_h_start+history_size-1)}{row}"
						")"
					)

				ws[sheet_cell]=tl

		if not inc_history:

			ws[sheet_cell]=supply_init+util_get_current_supply(
				asset[_KEY_HISTORY],
				date_min=date_min,
				date_max=date_max,
				the_trend=the_trend
			)

		# NEXT COLUMN (G) - THE PERFORMANCE (OPTIONAL)

		if atype==1 or atype==-1:

			col_pos=col_pos+1
			col_pos_ok=get_column_letter(col_pos)
			sheet_cell=f"{get_column_letter(col_pos)}{row}"

			if atype==1:

				# Uphill (CS - IS)
	
				ws[f"{col_pos_ok}{row}"]=(
					f"=SUM({sheet_cell_csup}-{sheet_cell_isup})"
				)

			if atype==-1:
	
				# Downhill (IS - CS)

				ws[f"{col_pos_ok}{row}"]=(
					f"=SUM({sheet_cell_isup}-{sheet_cell_csup})"
				)

		# NEXT COLUMN AND BEYOND - FULL HISTORY

		if not inc_history:
			continue

		col_pos=col_pos+1

		# history column index
		col_idx=-1

		print("\tReading history")

		for uid in asset[_KEY_HISTORY]:

			col_idx=col_idx+1

			record_date,valid_date=util_valid_date(
				asset[_KEY_HISTORY][uid].get(_KEY_DATE),
				date_min=date_min,
				date_max=date_max,
				fullres=True
			)
			if record_date is None:
				continue

			record_mod=util_valid_int(
				asset[_KEY_HISTORY][uid].get(_KEY_RECORD_MOD)
			)
			if record_mod is None:
				continue

			valid_trend=True
			if follow_trend and valid_date:
				valid_trend=util_filter_by_trend(
					record_mod,
					the_trend
				)

			record_comment=util_valid_str(
				asset[_KEY_HISTORY][uid].get(_KEY_COMMENT)
			)

			record_tag=util_valid_str(
				asset[_KEY_HISTORY][uid].get(_KEY_TAG),True
			)

			cell_comment=f"ID: {uid}"

			cell_comment=(
				f"{cell_comment}\n"
				f"{_TL_DATE[lang]}: {record_date}"
			)

			ignore_this=(
				(not valid_date) or
				(not valid_trend)
			)

			if ignore_this:
				cell_comment=(
					f"{cell_comment}\n\n"
					f"{_TL_IGNORED[lang]}: {record_mod}\n"
				)

				if not valid_date:
					cell_comment=(
						f"{cell_comment}\n"
						f"* {_ERR_TIMEFRAME[lang]}:"
					)
					if has_min:
						cell_comment=(
							f"{cell_comment}\n"
							f"- MIN: {date_min_str}"
						)
					if has_max:
						cell_comment=(
							f"{cell_comment}\n"
							f"- MAX: {date_max_str}"
						)

					cell_comment=f"{cell_comment}\n"

				if not valid_trend:
					cell_comment=(
						f"{cell_comment}\n"
						f"* {_ERR_TREND[lang]}\n"
						f"{_TL_TREND[lang]}"
					)
					if the_trend>0:
						cell_comment=f"{cell_comment}: {_TL_TREND_U[lang]}"
					if the_trend<0:
						cell_comment=f"{cell_comment}: {_TL_TREND_D[lang]}"

					cell_comment=f"{cell_comment}\n"

			if record_tag is not None:
				tl={
					_LANG_EN:"Tag",
					_LANG_ES:"Etiqueta"
				}[lang]
				cell_comment=(
					f"{cell_comment}\n"
					f"{tl}: {record_tag}"
				)

			if record_comment is not None:
				cell_comment=(
					f"{cell_comment}" "\n\n"
					f"{record_comment}"
				)

			print(
				"\t\t-",uid,
				record_date,
				record_mod
			)

			tgt_cell:Cell=ws[f"{get_column_letter(col_pos+col_idx)}{row}"]

			if ignore_this:
				record_mod=0

			tgt_cell.value=record_mod

			tgt_cell.comment=Comment(
				cell_comment,
				"SHLED",
				height=240,
				width=360
			)

	try:
		wb.save(path_output)
	except Exception as exc:
		print(_ExExErr,exc)
		return None

	return path_output

async def main(

		lang:str,
		path_base:Path,

		rdbc:AsyncIOMotorClient,
			rdbn:str,

		atype=0,
		inc_history:bool=False,

		date:Optional[datetime]=None,
			date_min:Optional[datetime]=None,
			date_max:Optional[datetime]=None,

		follow_trend:bool=False,

	)->Optional[Path]:

	print("Exporting assets to excel file")
	print("\tdate_min",date_min)
	print("\tdate_max",date_max)

	result_aq=await dbi_assets_AssetQuery(
		rdbc,rdbn,
		get_tag=True,
		get_value=True,
		get_history=True
	)
	if len(result_aq)==0:
		print(
			_ExExErr,
			"there are no assets"
		)
		return None

	return (
		await async_run_block(
			conversion_process,
			lang,path_base,
			result_aq,atype,inc_history,
			date,date_min,date_max,
			follow_trend
		)
	)

if __name__=="__main__":

	from asyncio import run as async_run
	from subprocess import run as sub_run

	from sys import (
		argv as sys_argv,
		exit as sys_exit,
		platform as sys_platform
	)

	rdbc=AsyncIOMotorClient()
	rdbn="test"

	all_assets=async_run(
		dbi_assets_AssetQuery(
			rdbc,rdbn,
			get_value=True,
			get_history=True
		)
	)

	if len(all_assets)==0:
		print(_ExExErr,"You have no assets")
		sys_exit(1)

	path_file=conversion_process(

		"es",Path(sys_argv[0]).parent,

		all_assets,

		atype=-1,
		inc_history=True,

		date_min_og=datetime(2025,2,19),
		date_max_og=datetime(2025,2,20),

		# history_bl_tags=["factura-itm"],

		follow_trend=True,
	)
	if path_file is None:
		print("Unable to create the excel file")
		sys_exit(1)

	on_windows=sys_platform.startswith("win")
	on_linux=sys_platform.startswith("linux")

	if on_linux or on_windows:

		program=""
		if on_linux:
			program="libreoffice"
		if on_windows:
			program="explorer"

		proc=sub_run(
			[
				program,
				f"{str(path_file)}"
			]
		)

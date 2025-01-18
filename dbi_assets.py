#!/usr/bin/python3.9

# import asyncio

import secrets

from typing import Mapping,Optional,Union

from motor.motor_asyncio import (
	AsyncIOMotorClient,
	AsyncIOMotorCursor,
	AsyncIOMotorCollection
)

# from pymongo.results import InsertOneResult
from pymongo.results import UpdateResult

# from dbi_accounts import (
# 	util_userid_to_backend,
# 	util_userid_from_backend,
# )

from internals import (
	util_rnow,
	util_valid_int,
	util_valid_str,
	# util_valid_date
)

from symbols_Any import (

	_ERR,
	_KEY_TAG,
	_KEY_SIGN,
	_KEY_DATE,
	_KEY_COMMENT,
)

from symbols_assets import (

	_COL_ASSETS,

	_KEY_ASSET,
	_KEY_NAME,
	_KEY_VALUE,
	_KEY_TOTAL,

	_KEY_HISTORY,
	_KEY_RECORD_UID,
	_KEY_RECORD_MOD,
)

def util_get_total_from_history(history:Mapping)->Mapping:
	if len(history)==0:
		return 0

	total=0
	for key in history:
		total=total+util_valid_int(
			history[key].get(_KEY_RECORD_MOD),
			fallback=0
		)

	return total

def util_calculate_total_in_asset(
		asset:Mapping,
		mutate:bool=False
	)->Union[bool,int]:

	if _KEY_HISTORY not in asset.keys():
		if not mutate:
			return 0
		return False

	if not isinstance(asset[_KEY_HISTORY],Mapping):
		if not mutate:
			return 0
		return False

	asset_total=util_get_total_from_history(asset[_KEY_HISTORY])
	if not mutate:
		return asset_total

	asset.update({_KEY_TOTAL:asset_total})

	return True

async def dbi_assets_CreateAsset(
		rdbc:AsyncIOMotorClient,name_db:str,
		asset_id:str,asset_name:str,
		asset_sign:str,
		asset_comment:Optional[str]=None,
		asset_tag:Optional[str]=None,
		asset_value:int=0,
		verblvl:int=2,
	)->Mapping:

	v=verblvl
	if verblvl not in range(0,3):
		v=2

	new_asset={
		"_id":asset_id,
		_KEY_NAME:asset_name,
		_KEY_SIGN:asset_sign,
		_KEY_VALUE:asset_value
	}

	if isinstance(asset_comment,str):
		if not len(asset_comment)==0:
			new_asset.update({_KEY_COMMENT:asset_comment})

	if isinstance(asset_tag,str):
		if not len(asset_tag)==0:
			new_asset.update({_KEY_TAG:asset_tag})

	try:
		tgtcol:AsyncIOMotorCollection=rdbc[name_db][_COL_ASSETS]
		await tgtcol.insert_one(new_asset)
	except Exception as exc:
		return {_ERR:f"{exc}"}

	if v==0:
		return {}

	if v==1:
		return {_KEY_ASSET:asset_id}

	new_asset.pop("_id")
	new_asset.update({
		_KEY_ASSET:asset_id,
	})
	return new_asset

async def dbi_assets_ChangeAssetMetadata(

		rdbc:AsyncIOMotorClient,

		name_db:str,
		asset_id:str,

		asset_name:Optional[str]=None,
		asset_tag:Optional[str]=None,
		asset_comment:Optional[str]=None,
		asset_value:Optional[int]=None,

		change_name:bool=False,
		change_tag:bool=False,
		change_comment:bool=False,
		change_value:bool=False

	)->Mapping:

	changes_set={}
	changes_unset=[]

	if change_name:
		name_ok=util_valid_str(asset_name)
		if name_ok:
			changes_set.update({_KEY_NAME:asset_name})
		if not name_ok:
			changes_set.update({_KEY_NAME:asset_id})

	if change_tag:
		tag_ok=util_valid_str(asset_tag)
		if tag_ok:
			changes_set.update({_KEY_TAG:asset_tag})
		if not tag_ok:
			# changes_unset.update({"$unset":{_KEY_TAG:1}})
			changes_unset.append(_KEY_TAG)

	if change_comment:
		comment_ok=util_valid_str(asset_comment)
		if comment_ok:
			changes_set.update({_KEY_COMMENT:asset_comment})
		if not comment_ok:
			changes_unset.append(_KEY_COMMENT)

	if change_value:
		value_ok=util_valid_str(asset_value)
		if value_ok:
			changes_set.update({_KEY_VALUE:asset_value})
		if not value_ok:
			changes_set.update({_KEY_VALUE:0})

	if len(changes_set)==0 and len(changes_unset)==0:
		return {_ERR:"Nothing to change"}

	aggr_pipeline=[{"$match":{"_id":asset_id}}]

	if len(changes_set)>0:
		aggr_pipeline.append({"$set":changes_set})

	if len(changes_unset)>0:
		aggr_pipeline.append({"$unset":changes_unset})

	# TODO: it should be dropped... ?
	# https://www.mongodb.com/docs/manual/reference/operator/aggregation/merge/

	print("\nAGGREGATION PIPELINE:",aggr_pipeline)

	aggr_pipeline.append(
		{
			"$merge":{
				"into":_COL_ASSETS,
				"whenMatched":"replace",
				"whenNotMatched":"insert"
			}
		}
	)

	# print("PIPELINE:",aggr_pipeline)

	try:
		col:AsyncIOMotorCollection=rdbc[name_db][_COL_ASSETS]
		cursor:AsyncIOMotorCursor=col.aggregate(aggr_pipeline)
		async for x in cursor:
			pass

	except Exception as exc:
		return {_ERR:f"{exc}"}

	# print(cursor)

	return {}

async def dbi_assets_AssetQuery(

		rdbc:AsyncIOMotorClient,name_db:str,

		asset_id:Optional[str]=None,
		asset_sign:Optional[str]=None,
		asset_tag:Optional[str]=None,

		get_sign:bool=False,
		get_tag:bool=False,
		get_comment:bool=False,
		get_total:bool=False,
		get_history:bool=False,
		get_value:bool=False,

	)->Union[Mapping,list]:

	only_one=isinstance(asset_id,str)

	spec_sign=isinstance(asset_sign,str)
	spec_tag=isinstance(asset_tag,str)

	find_match={}
	projection={"_id":1,_KEY_NAME:1}
	if isinstance(asset_id,str):
		find_match.update({"_id":asset_id})
	if spec_tag:
		find_match.update({_KEY_TAG:asset_tag})
		projection.update({_KEY_TAG:asset_tag})
	if spec_sign:
		find_match.update({_KEY_SIGN:asset_sign})
		projection.update({_KEY_SIGN:asset_sign})

	if get_sign and (not spec_sign):
		projection.update({_KEY_SIGN:1})
	if get_tag and (not spec_tag):
		projection.update({_KEY_TAG:1})
	if get_comment:
		projection.update({_KEY_COMMENT:1})
	if get_total or get_history:
		projection.update({_KEY_HISTORY:1})
	if get_value:
		projection.update({_KEY_VALUE:1})

	list_of_assets=[]
	try:
		tgtcol=rdbc[name_db][_COL_ASSETS]
		cursor:AsyncIOMotorCursor=tgtcol.aggregate([
			{"$match":find_match},
			{"$project":projection},
			{"$set":{_KEY_ASSET:"$_id","_id":"$$REMOVE"}}
		])
		async for asset in cursor:
			list_of_assets.append(asset)

	except Exception as exc:
		if only_one:
			return {_ERR:f"{exc}"}

		# NOTE: Never remove this
		print(exc)
		return []

	if get_total:
		for asset in list_of_assets:
			total=util_calculate_total_in_asset(asset)
			asset.update({_KEY_TOTAL:total})

	if only_one:
		return list_of_assets.pop()

	return list_of_assets

async def dbi_assets_DropAsset(
		rdbc:AsyncIOMotorClient,name_db:str,
		asset_id:str,outverb:int=2
	)->Mapping:

	v=outverb
	if outverb not in range(0,3):
		v=2

	result:Optional[Mapping]=None
	try:
		tgtcol:AsyncIOMotorCollection=rdbc[name_db][_COL_ASSETS]
		result=await tgtcol.find_one_and_delete(
			{"_id":asset_id}
		)
	except Exception as exc:
		return {_ERR:f"{exc}"}

	if v==0:
		return {}

	if v==1:
		return {_KEY_ASSET:asset_id}

	return result

# QUESTION: Have every record carry the value at the moment?

async def dbi_assets_History_AddRecord(
		rdbc:AsyncIOMotorClient,name_db:str,
		asset_id:str,
		record_sign:str,
		record_mod:int,
		record_tag:Optional[str]=None,
		record_comment:Optional[str]=None,
		vlevel:int=2
	)->Mapping:

	v=vlevel
	if vlevel not in range(0,3):
		v=2

	record_date=util_rnow()
	record_uid=secrets.token_hex(8)

	record_object={
		_KEY_DATE:record_date,
		_KEY_RECORD_MOD:record_mod,
		_KEY_SIGN:record_sign,
	}

	if isinstance(record_tag,str):
		record_object.update({_KEY_TAG:record_tag})

	if isinstance(record_comment,str):
		record_object.update({_KEY_COMMENT:record_comment})

	res:Optional[UpdateResult]=None

	try:
		col:AsyncIOMotorCollection=rdbc[name_db][_COL_ASSETS]
		res=await col.update_one(
			{ "_id" : asset_id } ,
			{"$set":{
					f"{_KEY_HISTORY}.{record_uid}": record_object
				}
			}
		)

	except Exception as exc:
		return {_ERR:f"{exc}"}

	if res.modified_count==0:
		return {_ERR:"???"}

	if v==0:
		return {}

	if v==1:
		return {
			_KEY_RECORD_UID:record_uid,
		}

	record_object.update({
		_KEY_RECORD_UID:record_uid,
	})

	return record_object

async def dbi_assets_History_GetSingleRecord(
		rdbc:AsyncIOMotorClient,name_db:str,
		asset_id:str,record_uid:str,
	)->Mapping:

	aggr_pipeline=[
		{
			"$match":{
				"_id":asset_id
			}
		},
		{
			"$project":{
				f"history.{record_uid}":1
			}
		},
		{
			"$set":{
				_KEY_ASSET:"$_id",
				"_id":"$$REMOVE"
			}
		}
	]

	result={}

	try:
		col:AsyncIOMotorCollection=rdbc[name_db][_COL_ASSETS]
		cursor:AsyncIOMotorCursor=col.aggregate(aggr_pipeline)
		async for doc in cursor:
			result.update(doc)
			break

	except Exception as exc:
		return {_ERR:f"{exc}"}

	if len(result)==0:
		return {_ERR:"Nothing was found"}

	if not isinstance(result.get(_KEY_HISTORY),Mapping):
		return {_ERR:"No history/records found in the asset"}

	if not isinstance(result[_KEY_HISTORY].get(record_uid),Mapping):
		return {_ERR:"The specified record was not found in the history"}

	the_record=result[_KEY_HISTORY][record_uid]

	the_record.update({_KEY_RECORD_UID:record_uid})

	return the_record

if __name__=="__main__":

	import asyncio

	from motor.motor_asyncio import AsyncIOMotorClient
	from openpyxl import Workbook
	from openpyxl.cell.cell import Cell
	from openpyxl.comments import Comment
	from openpyxl.worksheet.worksheet import Worksheet
	from pathlib import Path

	from internals import util_excel_dectocol,util_valid_date

	rdbc=AsyncIOMotorClient()

	all_assets=asyncio.run(
		dbi_assets_AssetQuery(
			rdbc,"my-inventory",
			get_history=True
		)
	)

	print(all_assets)

	test_file=Path("TEST.xlsx")
	if test_file.is_file():
		test_file.unlink()

	wb:Workbook=Workbook()
	ws:Worksheet=wb.active
	ws.title="Assets"
	ws.append(["ID",_KEY_NAME,_KEY_TOTAL])
	row=1
	for asset in all_assets:
		row=row+1

		asset_name=asset.get(_KEY_NAME)
		asset_id=asset.get("id")

		ws[f"A{row}"]=asset_id
		ws[f"B{row}"]=asset_name

		has_history=isinstance(asset.get(_KEY_HISTORY),Mapping)
		if not has_history:
			ws[f"C{row}"]=0
			continue

		history_len=len(asset[_KEY_HISTORY])
		if history_len==0:
			ws[f"C{row}"]=0
			continue

		col_start=4
		col_end=col_start+history_len-1

		ef=f"=SUM({util_excel_dectocol(col_start)}{row}:{util_excel_dectocol(col_end)}{row})"

		ws[f"C{row}"]=ef

		col_idx=-1

		for uid in asset[_KEY_HISTORY]:

			col_idx=col_idx+1

			record_mod=util_valid_int(
				asset[_KEY_HISTORY][uid].get(_KEY_RECORD_MOD)
			)

			record_comment=util_valid_str(
				asset[_KEY_HISTORY][uid].get(_KEY_COMMENT)
			)

			record_date=util_valid_date(
				asset[_KEY_HISTORY][uid].get(_KEY_DATE)
			)

			cell_comment=(
				f"UID:\n"
				f"{uid}"
			)

			if record_date is not None:
				cell_comment=(
					f"{cell_comment}\n"
					f"DATE:\n"
					f"{record_date}"
				)

			if record_comment is not None:
				cell_comment=(
					f"{cell_comment}" "\n\n"
					f"{record_comment}"
				)

			if record_mod is None:
				cell_comment=(
					f"{cell_comment}\n\n(WARNING)"
				)

			tgt_cell:Cell=ws[f"{util_excel_dectocol(col_start+col_idx)}{row}"]
			tgt_cell.value=record_mod
			tgt_cell.comment=Comment(
				cell_comment,
				"?",
				height=160,width=160
			)

	wb.save(test_file.name)

#!/usr/bin/python3.9

from asyncio import to_thread as async_run_block
# from datetime import (datetime,timedelta)
from pathlib import Path
from secrets import token_hex
from typing import Optional,Mapping

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from dbi_assets import dbi_assets_AssetQuery
from dbi_orders import dbi_orders_QueryOrders

from symbols_Any import (

	_DIR_TEMP,

	_LANG_EN,
	_LANG_ES,

	# _KEY_DATE,
	_KEY_TAG,
	_KEY_COMMENT,
)

from symbols_assets import (
	_COL_ASSETS,
	_KEY_ASSET,_KEY_NAME,
	_KEY_VALUE,_KEY_RECORD_MOD
)

from symbols_orders import (
	_KEY_ORDER_IS_FLIPPED,
	# _KEY_LOCKED_BY
)

from internals import (
	util_valid_bool,
	util_valid_int,
	util_valid_str,
)

_KEY_INC_TAGS="include-tags"
_KEY_AS_RAW="as-raw"

def asset_as_row(
		asset_id:str,
		doc_asset:Mapping,

		row:int,
		col_value:str,
		col_mod:str,
		# col_cap:str,

		include_tags:bool=False,
		flip_mods:bool=False,
	)->list:

	asset_name=doc_asset[_KEY_NAME]

	the_row=[asset_id,asset_name]

	if include_tags:
		asset_tag=doc_asset.get(_KEY_TAG)
		the_row.append(asset_tag)

	asset_value:Optional[int]=util_valid_int(
		doc_asset.get(_KEY_VALUE),
		fallback=0
	)
	the_row.append(asset_value)

	asset_mod:Optional[int]=util_valid_int(
		doc_asset.get(_KEY_RECORD_MOD),
		fallback=0
	)
	if flip_mods:
		the_row.append(-1*asset_mod)
	if not flip_mods:
		the_row.append(asset_mod)

	the_row.append(
		f"={col_value}{row}*{col_mod}{row}"
	)

	return the_row

def conversion_process(

		lang:str,
		path_base:Path,

		doc_order:Mapping,

		include_tags:bool=False,
		remove_facade:bool=False,

	)->Optional[Path]:

	path_output=path_base.joinpath(
		_DIR_TEMP
	).joinpath(
		f"order_x{token_hex(8)}.xlsx"
	)

	path_output.parent.mkdir(
		parents=True,
		exist_ok=True
	)

	col_headers=[

		{
			_LANG_EN:"Asset ID",
			_LANG_ES:"ID del activo"
		}[lang],

		{
			_LANG_EN:"Name",
			_LANG_ES:"Nombre"
		}[lang]
	]
	if include_tags:
		col_headers.append(
			{
				_LANG_EN:"Tag",
				_LANG_ES:"Etiqueta"
			}[lang]
		)

	col_headers.extend(
		[
			{
				_LANG_EN:"Value",
				_LANG_ES:"Valor"
			}[lang],
			{
				_LANG_EN:"Amount",
				_LANG_ES:"Cantidad"
			}[lang],
			"Cap."
		]
	)

	col_idx_cap=len(col_headers)
	col_idx_mod=col_idx_cap-1
	col_idx_value=col_idx_mod-1

	col_cap=get_column_letter(col_idx_cap)
	col_mod=get_column_letter(col_idx_mod)
	col_value=get_column_letter(col_idx_value)

	wb:Workbook=Workbook()
	ws:Worksheet=wb.active
	ws.title="SHLED_ORDER"

	row=1
	ws.append(col_headers)

	flip_mods=False
	if remove_facade:
		if util_valid_bool(
			doc_order.get(_KEY_ORDER_IS_FLIPPED),
			dval=False
		):
			flip_mods=True

	assets_inside=0

	for asset_id in doc_order[_COL_ASSETS]:

		assets_inside=assets_inside+1

		row=row+1

		ws.append(
			asset_as_row(
				asset_id,
				doc_order[_COL_ASSETS][asset_id],

				row,
				col_value,
				col_mod,

				include_tags=include_tags,
				flip_mods=flip_mods
			)
		)

	only_one=(assets_inside==1)

	if not only_one:

		ws[f"{get_column_letter(col_idx_value)}{row+2}"]={
			_LANG_EN:"Total value",
			_LANG_ES:"Valor total"
		}[lang]

		ws[f"{col_cap}{row+2}"]=(
			f"=SUM({col_cap}{2}:{col_cap}{row})"
		)

	order_comment=util_valid_str(
		doc_order.get(_KEY_COMMENT)
	)
	if order_comment is not None:
		tgt_cell:Cell=ws[f"A{row+2}"]
		tgt_cell.value="???"
		tgt_cell.comment=Comment(
			order_comment,"SHLED",
			height=240,width=360
		)

	try:
		wb.save(path_output)
	except Exception as exc:
		print(exc)
		return None

	return path_output

async def main(
		lang:str,path_base:Path,
		rdbc:AsyncIOMotorClient,rdbn:str,
		order_id:str,

		include_tags:bool=False,
		remove_facade:bool=False,

	)->Optional[Path]:

	# TODO: do some research on sqlite's FTS module ASAP

	# dbi_result=await dbi_orders_QueryOrders(
	doc_order=await dbi_orders_QueryOrders(

		rdbc,rdbn,
		order_id=order_id,

		include_assets=True,
		include_comment=True
	)
	# print("FOUND THIS",dbi_result)
	# if len(dbi_result)==0:
	# 	print("query result is empty")
	# 	return None

	# doc_order=dbi_result.pop()

	if not isinstance(
		doc_order.get(_COL_ASSETS),
		Mapping
	):
		print("No assets?",doc_order)
		return None

	if len(doc_order[_COL_ASSETS])==0:
		print("Assets subdoc empty?",doc_order)
		return None

	the_list=["0"]
	the_list.extend(
		doc_order[_COL_ASSETS].keys()
	)

	the_assets=await dbi_assets_AssetQuery(
		rdbc,rdbn,
		asset_id_list=the_list,
		get_value=True,
		get_tag=True
	)
	if len(the_assets)==0:
		print("Assets not found")
		return None

	for doc_asset in the_assets:
		asset_id=doc_asset[_KEY_ASSET]
		asset_name=doc_asset[_KEY_NAME]
		doc_order[_COL_ASSETS][asset_id].update({_KEY_NAME:asset_name})

	print("Order ready",doc_order)

	the_assets.clear()

	path_file=await async_run_block(
		conversion_process,
		lang,path_base,doc_order,
		include_tags,
		remove_facade
	)
	if not isinstance(path_file,Path):
		print("Failed to create the spreadsheet")
		return None

	return path_file

if __name__=="__main__":

	from asyncio import run as async_run
	from subprocess import run as sub_run

	from sys import (
		argv as sys_argv,
		exit as sys_exit,
		platform as sys_platform
	)

	rdbc=AsyncIOMotorClient()
	rdbn="tests"

	path_wdir=Path(sys_argv[0]).parent

	order_id=sys_argv[1]

	path_file=async_run(
		main(
			"en",path_wdir,
			rdbc,rdbn,
			order_id,
			remove_facade=False
		)
	)
	if path_file is None:
		print("Unable to create the excel file")
		sys_exit(1)

	on_windows=sys_platform.startswith("win")
	on_linux=sys_platform.startswith("linux")

	if on_linux or on_windows:

		program=""
		if on_linux:
			program="libreoffice"
		if on_windows:
			program="explorer"

		proc=sub_run(
			[
				program,
				f"{str(path_file)}"
			]
		)
